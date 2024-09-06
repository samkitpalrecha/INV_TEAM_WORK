from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import pandas as pd
import re
import openpyxl
from openpyxl.styles import Font

# define options //by this the browser window will remain open after your script finishes execution
options = Options()
options.add_experimental_option("detach", True)

# define the chroome webdriver that we are using for the webscraping purpose
driver = webdriver.Chrome(options = options)

# get the website which we want to scrape
website = driver.get('https://www.screener.in/company/TATAMOTORS/consolidated/')


#\\\\\Income Statement/////
search = driver.find_element(By.ID, "profit-loss")

#find the table
table = search.find_element(By.CSS_SELECTOR, "table.data-table.responsive-text-nowrap")

#click all buttons in the table
buttons = table.find_elements(By.CSS_SELECTOR, "button.button-plain")
for button in buttons:
   button.click()

# Extract table data row by row after clicking all buttons
table_rows = table.find_elements(By.TAG_NAME, 'tr')  # Locate all rows in the table

PNL = []

# Iterate over each row and extract cell data
for row in table_rows:
    # Locate all cells (td elements) in the row
    cells = row.find_elements(By.TAG_NAME, 'td')
    # Extract data from each cell and store
    row_data = [cell.text for cell in cells]
    PNL.append(row_data) 

PNL.pop(0)

df_PNL = pd.DataFrame(PNL)

# print(df_PNL)


#\\\\\Balance Sheet/////
search = driver.find_element(By.ID, "balance-sheet")

#find the table
table = search.find_element(By.CSS_SELECTOR, "table.data-table.responsive-text-nowrap")

#click all buttons in the table
buttons = table.find_elements(By.CSS_SELECTOR, "button.button-plain")
for button in buttons:
   button.click()

# Extract table data row by row after clicking all buttons
table_rows = table.find_elements(By.TAG_NAME, 'tr')  # Locate all rows in the table

BLST = []

# Iterate over each row and extract cell data
for row in table_rows:
    # Locate all cells (td elements) in the row
    cells = row.find_elements(By.TAG_NAME, 'td')
    # Extract data from each cell and print
    row_data = [cell.text for cell in cells]
    BLST.append(row_data) 

BLST.pop(0)

df_BLST = pd.DataFrame(BLST)

# print(df_BLST)

#\\\\\Cah Flow/////
search = driver.find_element(By.ID, "cash-flow")

#find the table
table = search.find_element(By.CSS_SELECTOR, "table.data-table.responsive-text-nowrap")

time.sleep(2)

#click all buttons in the table
buttons = table.find_elements(By.CSS_SELECTOR, "button.button-plain")
for button in buttons:
   button.click()

time.sleep(2)

# Extract table data row by row after clicking all buttons
table_rows = table.find_elements(By.TAG_NAME, 'tr')  # Locate all rows in the table

CF = []

# Iterate over each row and extract cell data
for row in table_rows:
    # Locate all cells (td elements) in the row
    cells = row.find_elements(By.TAG_NAME, 'td')
    # Extract data from each cell and print
    row_data = [cell.text for cell in cells]
    CF.append(row_data) 

CF.pop(0)

df_CF = pd.DataFrame(CF)

# print(df_CF)

driver.quit()

df = pd.concat([df_PNL, df_BLST, df_CF], ignore_index=True)
# print(df)

# Connect to an existing Excel file
excel_file_path = r'C:\Users\Administrator\.vscode\All_Codes\Fin_Statements_Template.xlsx'
workbook = openpyxl.load_workbook(excel_file_path)

# Duplicate the template sheet
new_sheet = workbook.copy_worksheet(workbook['Sheet1'])

# Rename the new sheet with the company symbol
new_sheet.title = 'TATAMOTORS'

#add the column names to the new_sheet
i=2
for column in df.columns[1:] :
  new_sheet.cell(row=1, column=i, value=column)
  i=i+1

# Write financial data to the new sheet
for index, row in df.iterrows():
    for j, value in enumerate(row, start=1):
        new_sheet.cell(row=index+2, column=j, value=value if pd.notnull(value) else None)

#add rows between the statements
new_sheet.insert_rows(28)
new_sheet.insert_rows(29)
new_sheet.insert_rows(64)
new_sheet.insert_rows(65)

# #freeze the 1st row
# Freeze the first column
new_sheet.freeze_panes = 'B2'

new_sheet.cell(row=1, column=1, value='TATAMOTORS')
for i in range(2,10):
    new_sheet.cell(row=1, column=i, value='Mar 201'+str(i))
for i in range(10,14):    
    new_sheet.cell(row=1, column=i, value='Mar '+str(2010 + int(i)))
new_sheet.cell(row=1, column=14, value='TTM/Sept 2023')

# Iterate through the cells in the first row and set the font to bold
for cell in new_sheet[1]:
    cell.font = Font(bold=True)

# Specify the range of cells containing the values ending with '%'
cell_range = 'A1:T100'  # Example range

# Iterate through rows and cells to remove '%' symbol
for row in new_sheet.iter_rows(min_row=2, max_row=100, min_col=2, max_col=20):
    for cell in row:
        if isinstance(cell.value, str) and cell.value.endswith('%'):
                cell.value = cell.value[:-1]  # Remove the last character, which is '%'

# Iterate through all cells in the sheet
for row in new_sheet.iter_rows(min_row=2, max_row=100, min_col=2, max_col=20):
    for cell in row:
        # Check if the cell value is a numeric string
        if isinstance(cell.value, str) and cell.value.replace('.', '', 1).lstrip('-').isdigit():
            # Convert the numeric string to a number
            cell.value = float(cell.value)

search = driver.get("https://in.tradingview.com/symbols/NSE-TATAMOTORS/")

beta = driver.find_element(By.CLASS_NAME, "apply-overflow-tooltip value-GgmpMpKr")

new_sheet.cell(row=100, column=1, value='Beta')
new_sheet.cell(row=100, column=2, value=beta.text)

# Set the font to bold for cells in the first column
for row in new_sheet.iter_rows(min_row=2, max_row=new_sheet.max_row, min_col=1, max_col=1):
    for cell in row:
        cell.font = Font(bold=True)

driver.quit()        

# Save the changes back to the Excel file
workbook.save(excel_file_path)
# Close the workbook
workbook.close()