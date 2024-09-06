from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import pandas as pd
import re
import openpyxl
from openpyxl.styles import Font

# define options //by this the browser window will remain open after your script finishes execution
options = Options()
options.add_experimental_option("detach", True)

# define the chroome webdriver that we are using for the webscraping purpose
driver = webdriver.Chrome(options = options)

search = driver.get("https://in.tradingview.com/symbols/NSE-TATAMOTORS/")

time.sleep(2)

beta = driver.find_element(By.CSS_SELECTOR, "#js-category-content > div.js-symbol-page-tab-overview-root > div > section > div:nth-child(2) > div.container-GRoarMHL > div:nth-child(8) > div.wrapper-GgmpMpKr > div.apply-overflow-tooltip.value-GgmpMpKr")

time.sleep(2)

excel_file_path = r'C:\Users\Administrator\.vscode\All_Codes\Fin_Statements_Template.xlsx'
workbook = openpyxl.load_workbook(excel_file_path)

sheet = workbook['TATAMOTORS']

sheet.cell(row=100, column=1, value='Beta')
sheet.cell(row=100, column=2, value=beta.text)

# Set the font to bold for cells in the first column
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
    for cell in row:
        cell.font = Font(bold=True)

# Save the changes back to the Excel file
workbook.save(excel_file_path)
# Close the workbook
workbook.close()        

driver.quit()